"""
Campaign Routes — Send email/WhatsApp campaigns, retrieve logs, update lead outcomes,
schedule Google Calendar meetings, and manage meeting records.
All endpoints are protected and scoped to the current user.
"""

from fastapi import APIRouter, Depends, HTTPException, Request
from pydantic import BaseModel
from typing import Optional, List
import re
import aiosqlite
from datetime import datetime

from backend.database import get_db
from backend.models.campaign import CampaignCreate, CampaignResponse
from backend.middleware.auth_middleware import get_current_user, require_admin
from backend.services.lead_service import get_lead_by_id, update_lead
from backend.services.user_service import get_user_by_id
from backend.services.campaign_service import (
    create_campaign,
    update_campaign_status,
    get_campaigns_for_user,
    get_campaign_stats,
    get_daily_email_usage,
    check_email_quota,
    get_campaign_counts_by_lead,
    get_daily_email_limit,
    get_office_hours,
    update_office_hours,
    validate_meeting_slot,
)
from backend.services.email_service import send_email, is_gmail_configured, update_gmail_config, GMAIL_CONFIG
from backend.services.whatsapp_service import send_whatsapp, is_interakt_configured, update_interakt_config, INTERAKT_CONFIG
from backend.services.google_calendar_service import (
    schedule_meeting,
    is_gcal_configured,
    update_gcal_config,
    GCAL_CONFIG,
)
from backend.services.imap_service import (
    GMAIL_FETCH_CONFIG,
    is_gmail_fetch_configured,
    update_gmail_fetch_config,
)
from backend.default_templates import POSH_DEFAULT_TEMPLATE, CONTACT_US_DEFAULT_TEMPLATE

router = APIRouter(prefix="/api/campaigns", tags=["Campaigns"])


# ──────────────────────────────────────────────
#  Send Email Campaign  (supports HTML body)
# ──────────────────────────────────────────────
@router.post("/email")
async def send_email_campaign(
    campaign_data: CampaignCreate,
    request: Request,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Send an email campaign to a lead.
    Requires the lead to have an email address.
    Pass is_html=true in the body to render the message as HTML.
    On success, automatically updates lead status from 'new' → 'contacted'.
    """
    if campaign_data.campaign_type != "email":
        raise HTTPException(status_code=400, detail="Campaign type must be 'email'")

    # ── Daily email quota check ──
    if not await check_email_quota(db):
        usage = await get_daily_email_usage(db, current_user["id"])
        raise HTTPException(
            status_code=429,
            detail=f"Daily email campaign limit ({usage['daily_limit']}) reached. Try again tomorrow.",
        )

    lead = await get_lead_by_id(db, campaign_data.lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    if not lead.get("email"):
        raise HTTPException(status_code=400, detail="This lead does not have an email address")

    # Create campaign record (pending)
    campaign = await create_campaign(
        db=db,
        lead_id=campaign_data.lead_id,
        owner_id=current_user["id"],
        campaign_type="email",
        subject=campaign_data.subject,
        message=campaign_data.message,
        status="pending",
    )

    # Build CC list: admin default CCs (outgoing scope) + frontend-supplied CCs
    cc_list = []
    cursor = await db.execute("SELECT email FROM cc_emails WHERE scope_outgoing = 1")
    admin_ccs = [row["email"] for row in await cursor.fetchall()]
    cc_list.extend(admin_ccs)
    if campaign_data.cc_emails:
        cc_list.extend(campaign_data.cc_emails)
    # Deduplicate
    cc_list = list(dict.fromkeys([e.strip().lower() for e in cc_list if e and e.strip()]))

    # Attempt to send email — support HTML flag from frontend
    # Auto-generate booking URL for the "Book Appointment" button
    booking_url = None
    if getattr(campaign_data, "include_booking", True):
        try:
            from backend.services.lead_service import generate_booking_token
            token = await generate_booking_token(db, campaign_data.lead_id)
            # Build full absolute booking URL from the request origin
            origin = str(request.base_url).rstrip("/")
            booking_url = f"{origin}/book-meeting?lead_id={campaign_data.lead_id}&token={token}"
        except Exception as e:
            print(f"[EMAIL] Warning: Could not generate booking URL: {e}")

    is_html = getattr(campaign_data, "is_html", False) or False

    # Build tracking pixel URL for email open tracking
    tracking_url = f"{origin}/api/track/open/{campaign['id']}"

    result = send_email(
        to_email=lead["email"],
        subject=campaign_data.subject or "Message from Lead Manager",
        body=campaign_data.message,
        html=is_html,
        cc_emails=cc_list if cc_list else None,
        booking_url=booking_url,
        tracking_pixel_url=tracking_url,
    )

    # Update campaign status based on result
    if result["success"]:
        campaign = await update_campaign_status(db, campaign["id"], "sent")
        # Auto-update lead status from "new" → "contacted" only after successful send
        if lead.get("status") == "new":
            await update_lead(db=db, lead_id=lead["id"], status="contacted")
    else:
        campaign = await update_campaign_status(db, campaign["id"], "failed", result["message"])

    return {
        "campaign": campaign,
        "send_result": result,
    }


# ──────────────────────────────────────────────
#  Send WhatsApp Campaign
# ──────────────────────────────────────────────
@router.post("/whatsapp")
async def send_whatsapp_campaign(
    campaign_data: CampaignCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Send a WhatsApp campaign to a lead.
    Requires the lead to have a phone number.
    On success, automatically updates lead status from 'new' → 'contacted'.
    """
    if campaign_data.campaign_type != "whatsapp":
        raise HTTPException(status_code=400, detail="Campaign type must be 'whatsapp'")

    lead = await get_lead_by_id(db, campaign_data.lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    if not lead.get("phone"):
        raise HTTPException(status_code=400, detail="This lead does not have a phone number")

    # Create campaign record
    campaign = await create_campaign(
        db=db,
        lead_id=campaign_data.lead_id,
        owner_id=current_user["id"],
        campaign_type="whatsapp",
        subject=None,
        message=campaign_data.message,
        status="pending",
    )

    # Attempt to send WhatsApp
    # Note: campaign_data.message contains the selected template code_name
    result = send_whatsapp(
        to_phone=lead["phone"],
        template_name=campaign_data.message,
    )

    if result["success"]:
        campaign = await update_campaign_status(db, campaign["id"], "sent")
        # Auto-update lead status from "new" → "contacted" only after successful send
        if lead.get("status") == "new":
            await update_lead(db=db, lead_id=lead["id"], status="contacted")
    else:
        campaign = await update_campaign_status(db, campaign["id"], "failed", result["message"])

    return {
        "campaign": campaign,
        "send_result": result,
    }


# ──────────────────────────────────────────────
#  List Campaigns
# ──────────────────────────────────────────────
@router.get("")
async def list_campaigns(
    campaign_type: str = None,
    campaign_status: str = None,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """List campaigns. Salesperson sees own; Admin sees all.
    Supports optional campaign_type and campaign_status filters."""
    campaigns = await get_campaigns_for_user(
        db=db,
        user_id=current_user["id"],
        role=current_user["role"],
        campaign_type=campaign_type,
        campaign_status=campaign_status,
    )
    return campaigns


# ──────────────────────────────────────────────
#  Download Campaigns as Excel
# ──────────────────────────────────────────────
@router.get("/download")
async def download_campaigns_excel(
    campaign_type: str = None,
    campaign_status: str = None,
    lead_status: str = None,
    search: str = None,
    date: str = None,
    email_open: str = None,
    limit: int = 0,
    offset: int = 0,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Download campaigns as a structured Excel file.
    Columns: Company Name, Contact Person Name, Email, Phone,
             Type, Subject, Campaign Status, Lead Status, Sent At, Sent By, Date.
    Supports campaign_type, campaign_status, lead_status, search, date, and email_open filters.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from datetime import datetime as dt
    from html import unescape

    # Strip HTML tags for clean plain-text in Excel
    def strip_html(html_str):
        if not html_str:
            return ""
        clean = re.sub(r"<[^>]+>", "", html_str)
        clean = unescape(clean)
        clean = re.sub(r"\s+", " ", clean).strip()
        return clean

    # Build query — same JOIN pattern as get_campaigns_for_user()
    query = """
        SELECT c.*, l.company_name, l.contact_name, l.email as lead_email,
               l.phone as lead_phone, l.status as lead_status, l.created_at as lead_created_at,
               u.full_name as sender_name
        FROM campaigns c
        LEFT JOIN leads l ON c.lead_id = l.id
        LEFT JOIN users u ON c.owner_id = u.id
    """
    params = []
    conditions = []

    # Ownership filter
    if current_user["role"] != "admin":
        conditions.append("c.owner_id = ?")
        params.append(current_user["id"])

    if campaign_type:
        conditions.append("c.campaign_type = ?")
        params.append(campaign_type)

    if campaign_status:
        conditions.append("c.status = ?")
        params.append(campaign_status)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY c.created_at DESC"

    cursor = await db.execute(query, params)
    rows = [dict(row) for row in await cursor.fetchall()]

    # Apply client-side-equivalent filters server-side
    if lead_status:
        rows = [r for r in rows if r.get("lead_status") == lead_status]

    if search:
        search_lower = search.lower()
        rows = [
            r for r in rows
            if search_lower in " ".join([
                r.get("company_name") or "",
                r.get("contact_name") or "",
                r.get("lead_email") or "",
                r.get("lead_phone") or "",
                r.get("subject") or "",
                r.get("message") or "",
                r.get("sender_name") or "",
            ]).lower()
        ]

    if date:
        rows = [
            r for r in rows
            if (r.get("sent_at") or r.get("created_at") or "")[:10] == date
        ]

    if email_open:
        if email_open == "opened":
            rows = [r for r in rows if r.get("campaign_type") == "email" and r.get("email_opened") == 1]
        elif email_open == "sent":
            rows = [r for r in rows if r.get("campaign_type") == "email" and r.get("status") == "sent" and not r.get("email_opened")]

    # Apply pagination limit/offset
    if limit and limit > 0:
        rows = rows[offset:offset + limit]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaigns"

    # Header styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers_list = [
        "Company Name", "Contact Person Name", "Email", "Phone",
        "Type", "Subject", "Campaign Status", "Lead Status",
        "Sent At", "Sent By", "Date",
    ]
    ws.append(headers_list)

    # Apply header styling
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Status label mappings
    status_labels = {
        "new": "New", "contacted": "Contacted", "won": "Approved",
        "lost": "Lost", "sent": "Sent", "failed": "Failed", "pending": "Pending",
    }
    type_labels = {"email": "Email", "whatsapp": "WhatsApp", "website": "Website"}

    # Data styling
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center")
    alt_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")

    for row_idx, row_data in enumerate(rows, start=2):
        # Format sent_at
        sent_at_val = ""
        if row_data.get("sent_at"):
            try:
                parsed = dt.fromisoformat(row_data["sent_at"].replace("Z", "+00:00"))
                sent_at_val = parsed.strftime("%d-%b-%Y %H:%M")
            except (ValueError, AttributeError):
                sent_at_val = str(row_data["sent_at"])[:16]

        # Format lead created_at (Date column)
        date_val = ""
        if row_data.get("lead_created_at"):
            try:
                parsed = dt.fromisoformat(row_data["lead_created_at"].replace("Z", "+00:00"))
                date_val = parsed.strftime("%d-%b-%Y")
            except (ValueError, AttributeError):
                date_val = str(row_data["lead_created_at"])[:10]

        data_row = [
            row_data.get("company_name") or "",
            row_data.get("contact_name") or "",
            row_data.get("lead_email") or "",
            row_data.get("lead_phone") or "",
            type_labels.get(row_data.get("campaign_type", ""), row_data.get("campaign_type", "")),
            strip_html(row_data.get("subject") or ""),
            status_labels.get(row_data.get("status", ""), row_data.get("status", "")),
            status_labels.get(row_data.get("lead_status", ""), row_data.get("lead_status", "")),
            sent_at_val,
            row_data.get("sender_name") or "",
            date_val,
        ]
        ws.append(data_row)

        # Apply data styling with alternating row colors
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    today = dt.now().strftime("%Y-%m-%d")
    resp_headers = {
        "Content-Disposition": f'attachment; filename="campaigns_export_{today}.xlsx"'
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


# ──────────────────────────────────────────────
#  Campaign Statistics (for dashboard)
# ──────────────────────────────────────────────
@router.get("/stats")
async def campaign_stats(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get campaign statistics for the dashboard."""
    stats = await get_campaign_stats(db, current_user["id"], current_user["role"])
    return stats


# ──────────────────────────────────────────────
#  Daily Email Usage (for Usage page)
# ──────────────────────────────────────────────
@router.get("/usage")
async def email_usage(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Get daily email campaign usage stats.
    Returns global usage (all users) and current user's personal usage.
    Accessible by both admin and salesperson.
    """
    usage = await get_daily_email_usage(db, user_id=current_user["id"])
    usage["user_name"] = current_user.get("full_name", current_user.get("user_id", ""))
    return usage


# ──────────────────────────────────────────────
#  Email Open Rate / Conversion Rate
# ──────────────────────────────────────────────
@router.get("/open-rate")
async def email_open_rate(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Get email open rate / conversion rate stats.
    Each user sees their own stats. Admin also sees per-user breakdown.
    """
    is_admin = current_user["role"] == "admin"
    user_id = current_user["id"]

    # Current user's stats
    if is_admin:
        cursor = await db.execute(
            "SELECT COUNT(*) as count FROM campaigns WHERE campaign_type = 'email' AND status = 'sent'"
        )
    else:
        cursor = await db.execute(
            "SELECT COUNT(*) as count FROM campaigns WHERE campaign_type = 'email' AND status = 'sent' AND owner_id = ?",
            (user_id,),
        )
    total_row = await cursor.fetchone()
    total_sent = total_row["count"] if total_row else 0

    if is_admin:
        cursor = await db.execute(
            "SELECT COUNT(*) as count FROM campaigns WHERE campaign_type = 'email' AND status = 'sent' AND email_opened = 1"
        )
    else:
        cursor = await db.execute(
            "SELECT COUNT(*) as count FROM campaigns WHERE campaign_type = 'email' AND status = 'sent' AND email_opened = 1 AND owner_id = ?",
            (user_id,),
        )
    opened_row = await cursor.fetchone()
    total_opened = opened_row["count"] if opened_row else 0

    open_rate = round((total_opened / total_sent * 100), 1) if total_sent > 0 else 0

    result = {
        "total_sent": total_sent,
        "total_opened": total_opened,
        "open_rate": open_rate,
    }

    # Admin: per-user breakdown
    if is_admin:
        cursor = await db.execute(
            """
            SELECT u.id as user_id, u.full_name,
                   COUNT(CASE WHEN c.status = 'sent' THEN 1 END) as sent,
                   COUNT(CASE WHEN c.status = 'sent' AND c.email_opened = 1 THEN 1 END) as opened
            FROM users u
            LEFT JOIN campaigns c ON c.owner_id = u.id AND c.campaign_type = 'email'
            WHERE u.is_active = 1
            GROUP BY u.id, u.full_name
            ORDER BY sent DESC
            """
        )
        per_user = []
        for row in await cursor.fetchall():
            sent = row["sent"] or 0
            opened = row["opened"] or 0
            per_user.append({
                "user_id": row["user_id"],
                "user_name": row["full_name"],
                "sent": sent,
                "opened": opened,
                "rate": round((opened / sent * 100), 1) if sent > 0 else 0,
            })
        result["per_user"] = per_user

    return result


# ──────────────────────────────────────────────
#  Campaign Counts per Lead (for dynamic buttons)
# ──────────────────────────────────────────────
@router.get("/lead-counts")
async def campaign_lead_counts(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Get the number of successfully sent campaigns per lead, grouped by type.
    Used by the leads table to render dynamic counter buttons.
    """
    counts = await get_campaign_counts_by_lead(
        db, current_user["id"], current_user["role"]
    )
    return counts


# ──────────────────────────────────────────────
#  Admin: Set Daily Email Campaign Limit
# ──────────────────────────────────────────────
class DailyLimitUpdate(BaseModel):
    limit: int

@router.post("/settings/daily-email-limit")
async def set_daily_email_limit(
    body: DailyLimitUpdate,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """
    Set the daily email campaign send limit (admin only).
    Persisted in system_settings.
    """
    if body.limit < 1:
        raise HTTPException(status_code=400, detail="Limit must be at least 1")

    await db.execute(
        "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('daily_email_limit', ?)",
        (str(body.limit),),
    )
    await db.commit()
    return {"message": f"Daily email limit updated to {body.limit}", "limit": body.limit}


@router.get("/settings/daily-email-limit")
async def get_daily_limit_setting(
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """Get the current daily email campaign limit (admin only)."""
    limit = await get_daily_email_limit(db)
    return {"limit": limit}


# ──────────────────────────────────────────────
#  Update Lead Outcome (Lost / Approved/Won)
# ──────────────────────────────────────────────
class LeadOutcomeRequest(BaseModel):
    outcome: str  # "lost" or "won"

@router.put("/lead/{lead_id}/outcome")
async def update_lead_outcome(
    lead_id: int,
    body: LeadOutcomeRequest,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Update a contacted lead's outcome to 'won' (Approved) or 'lost'.
    Only allowed when lead is in 'contacted' status.
    Called from the Campaigns page when a salesperson reviews lead feedback.
    """
    if body.outcome not in ("won", "lost"):
        raise HTTPException(status_code=400, detail="Outcome must be 'won' or 'lost'")

    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    updated = await update_lead(db=db, lead_id=lead_id, status=body.outcome)
    return updated


# ──────────────────────────────────────────────
#  Schedule Google Meet Meeting  +  Save to DB
# ──────────────────────────────────────────────
class MeetingRequest(BaseModel):
    lead_id: int
    title: str
    description: Optional[str] = ""
    start_datetime: str   # "YYYY-MM-DDTHH:MM:SS"
    duration_minutes: Optional[int] = 60
    attendee_email: Optional[str] = None
    cc_emails: Optional[List[str]] = None

@router.post("/schedule-meeting")
async def schedule_google_meeting(
    body: MeetingRequest,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Schedule a Google Meet meeting on the admin's Google Calendar.
    Saves meeting details to the local database for later retrieval / editing.
    Called after a lead is marked as 'Approved'.
    Returns a Google Meet link.
    """
    lead = await get_lead_by_id(db, body.lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # ── Validate meeting slot (office hours + overlap + buffer) ──
    # Check is per-user: only this user's meetings are considered
    validation = await validate_meeting_slot(
        db,
        owner_id=current_user["id"],
        start_datetime=body.start_datetime,
        duration_minutes=body.duration_minutes or 60,
        exclude_lead_id=body.lead_id,  # for edit/upsert case
    )
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail=validation["reason"])

    # Derive salesperson email from the current user's record
    sp_user = await get_user_by_id(db, current_user["id"])
    salesperson_email = sp_user.get("email", "") if sp_user else ""

    # Build CC list for meeting: admin default CCs (meetings scope) + frontend-supplied CCs
    meeting_cc_list = []
    cursor = await db.execute("SELECT email FROM cc_emails WHERE scope_meetings = 1")
    admin_meeting_ccs = [row["email"] for row in await cursor.fetchall()]
    meeting_cc_list.extend(admin_meeting_ccs)
    if body.cc_emails:
        meeting_cc_list.extend(body.cc_emails)
    # Deduplicate
    meeting_cc_list = list(dict.fromkeys([e.strip().lower() for e in meeting_cc_list if e and e.strip()]))

    result = schedule_meeting(
        title=body.title,
        description=body.description or f"Meeting with {lead.get('contact_name') or lead.get('company_name', 'Client')}",
        start_datetime=body.start_datetime,
        duration_minutes=body.duration_minutes or 60,
        attendee_email=body.attendee_email or lead.get("email"),
        salesperson_email=salesperson_email or None,
        cc_emails=meeting_cc_list if meeting_cc_list else None,
    )

    if not result["success"]:
        raise HTTPException(status_code=503, detail=result["message"])

    # Upsert meeting record in local DB (INSERT OR REPLACE by lead_id)
    now = datetime.utcnow().isoformat()
    await db.execute(
        """
        INSERT INTO meetings
            (lead_id, owner_id, title, description, start_datetime, duration_minutes,
             attendee_email, salesperson_email, event_id, event_link, meet_link, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(lead_id) DO UPDATE SET
            title=excluded.title,
            description=excluded.description,
            start_datetime=excluded.start_datetime,
            duration_minutes=excluded.duration_minutes,
            attendee_email=excluded.attendee_email,
            salesperson_email=excluded.salesperson_email,
            event_id=excluded.event_id,
            event_link=excluded.event_link,
            meet_link=excluded.meet_link,
            updated_at=excluded.updated_at
        """,
        (
            body.lead_id,
            current_user["id"],
            body.title,
            body.description or "",
            body.start_datetime,
            body.duration_minutes or 60,
            body.attendee_email or lead.get("email", ""),
            salesperson_email or "",
            result.get("event_id", ""),
            result.get("event_link", ""),
            result.get("meet_link", ""),
            now,
            now,
        ),
    )
    await db.commit()

    return result


# ──────────────────────────────────────────────
#  Get Scheduled Meeting for a Lead
# ──────────────────────────────────────────────
@router.get("/meeting/{lead_id}")
async def get_meeting(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get the scheduled meeting details for a specific lead."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    cursor = await db.execute("SELECT * FROM meetings WHERE lead_id = ?", (lead_id,))
    row = await cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="No meeting scheduled for this lead")

    return dict(row)


# ──────────────────────────────────────────────
#  Delete / Cancel a Scheduled Meeting
# ──────────────────────────────────────────────
@router.delete("/meeting/{lead_id}")
async def cancel_meeting(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Cancel/delete a scheduled meeting for a lead.
    Removes the local DB record. The Google Calendar event must be
    deleted manually (or use event_link to open it in Google Calendar).
    """
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    cursor = await db.execute("DELETE FROM meetings WHERE lead_id = ?", (lead_id,))
    await db.commit()

    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="No meeting found for this lead")

    return {"message": "Meeting cancelled successfully"}


# ──────────────────────────────────────────────
#  Check if a lead has a scheduled meeting
# ──────────────────────────────────────────────
@router.get("/meeting-status")
async def get_meeting_statuses(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Return a list of lead_ids that have scheduled meetings for this user.
    Used by the campaigns table to decide which action button to show.
    """
    if current_user["role"] == "admin":
        cursor = await db.execute("SELECT lead_id FROM meetings")
    else:
        cursor = await db.execute(
            "SELECT lead_id FROM meetings WHERE owner_id = ?", (current_user["id"],)
        )
    rows = await cursor.fetchall()
    return [row["lead_id"] for row in rows]


# ──────────────────────────────────────────────
#  Settings: Gmail API Config (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/gmail")
async def get_gmail_settings(admin: dict = Depends(require_admin)):
    """Get current Gmail API configuration (admin only)."""
    return {
        "client_id": GMAIL_CONFIG["client_id"],
        "client_secret": "***" if GMAIL_CONFIG["client_secret"] else "",
        "refresh_token": "***" if GMAIL_CONFIG["refresh_token"] else "",
        "from_email": GMAIL_CONFIG["from_email"],
        "from_name": GMAIL_CONFIG["from_name"],
        "configured": is_gmail_configured(),
    }


@router.post("/settings/gmail")
async def save_gmail_settings(
    config: dict,
    admin: dict = Depends(require_admin),
):
    """Update Gmail API configuration (admin only)."""
    update_gmail_config(config)
    return {"message": "Gmail API settings updated", "configured": is_gmail_configured()}


# ──────────────────────────────────────────────
#  Settings: Interakt WhatsApp Config (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/interakt")
async def get_interakt_settings(admin: dict = Depends(require_admin)):
    """Get current Interakt configuration (admin only)."""
    api_key = INTERAKT_CONFIG.get("api_key", "")
    return {
        "api_key": (api_key[:8] + "***") if api_key else "",
        "template_name": INTERAKT_CONFIG.get("template_name", ""),
        "language_code": INTERAKT_CONFIG.get("language_code", "en"),
        "configured": is_interakt_configured(),
    }


@router.post("/settings/interakt")
async def save_interakt_settings(
    config: dict,
    admin: dict = Depends(require_admin),
):
    """Update Interakt configuration (admin only)."""
    # Don't overwrite masked values
    if config.get("api_key", "").endswith("***"):
        config.pop("api_key", None)
    update_interakt_config(config)
    return {"message": "Interakt settings updated", "configured": is_interakt_configured()}


# ──────────────────────────────────────────────
#  Settings: WhatsApp Templates (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/whatsapp-templates")
async def get_whatsapp_templates(db: aiosqlite.Connection = Depends(get_db)):
    """Get all WhatsApp templates."""
    cursor = await db.execute("SELECT * FROM whatsapp_templates ORDER BY created_at DESC")
    rows = await cursor.fetchall()
    return [dict(row) for row in rows]


class WhatsAppTemplateCreate(BaseModel):
    name: str
    code_name: str


@router.post("/settings/whatsapp-templates")
async def create_whatsapp_template(
    body: WhatsAppTemplateCreate,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """Create a new WhatsApp template."""
    if not body.name or not body.code_name:
        raise HTTPException(status_code=400, detail="Name and code_name are required")
        
    await db.execute(
        "INSERT INTO whatsapp_templates (name, code_name) VALUES (?, ?)",
        (body.name, body.code_name)
    )
    await db.commit()
    return {"message": "Template created successfully"}


@router.delete("/settings/whatsapp-templates/{template_id}")
async def delete_whatsapp_template(
    template_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """Delete a WhatsApp template."""
    cursor = await db.execute("DELETE FROM whatsapp_templates WHERE id = ?", (template_id,))
    await db.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted successfully"}


# ──────────────────────────────────────────────
#  Settings: Email HTML Templates (User-scoped)
# ──────────────────────────────────────────────
@router.get("/settings/email-templates")
async def get_email_templates(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Get email templates visible to the current user.
    Returns:
      - Admin/system templates (owner_id IS NULL) — marked is_default=true
      - Current user's private templates (owner_id = current_user.id) — marked is_default=false
    If the user has forked an admin template, the original is excluded
    and replaced by the user's fork.
    """
    user_id = current_user["id"]

    # 1. Get user's private templates (including forks)
    cursor = await db.execute(
        "SELECT * FROM email_templates WHERE owner_id = ? ORDER BY COALESCE(updated_at, created_at) DESC",
        (user_id,),
    )
    user_templates = [dict(row) for row in await cursor.fetchall()]

    # 2. Collect IDs of admin templates the user has forked
    forked_admin_ids = {t["forked_from"] for t in user_templates if t.get("forked_from")}

    # 3. Get admin templates, excluding ones the user has forked
    cursor = await db.execute(
        "SELECT * FROM email_templates WHERE owner_id IS NULL ORDER BY created_at DESC"
    )
    admin_templates = [dict(row) for row in await cursor.fetchall()]
    admin_templates = [t for t in admin_templates if t["id"] not in forked_admin_ids]

    # 4. Mark and combine
    for t in admin_templates:
        t["is_default"] = True
    for t in user_templates:
        t["is_default"] = False

    return admin_templates + user_templates


class EmailTemplateCreate(BaseModel):
    name: str
    subject: Optional[str] = None
    html_body: str
    is_default: Optional[bool] = False  # Admin can create system templates


@router.post("/settings/email-templates")
async def create_email_template(
    body: EmailTemplateCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Create a new Email HTML template.
    - Admin can create system templates (owner_id=NULL) by passing is_default=true.
    - Any user can create a private template (owner_id=current_user.id).
    """
    if not body.name or not body.html_body:
        raise HTTPException(status_code=400, detail="Name and HTML body are required")

    now = datetime.utcnow().isoformat()

    # Admin creating a system/default template
    if body.is_default and current_user["role"] == "admin":
        owner_id = None
    else:
        owner_id = current_user["id"]

    cursor = await db.execute(
        """INSERT INTO email_templates (name, subject, html_body, owner_id, forked_from, created_at, updated_at)
           VALUES (?, ?, ?, ?, NULL, ?, ?)""",
        (body.name, body.subject, body.html_body, owner_id, now, now),
    )
    await db.commit()
    new_id = cursor.lastrowid

    # Return the created template
    cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (new_id,))
    row = await cursor.fetchone()
    result = dict(row)
    result["is_default"] = owner_id is None
    return result


class EmailTemplateUpdate(BaseModel):
    name: str
    subject: Optional[str] = None
    html_body: str


@router.put("/settings/email-templates/{template_id}")
async def update_email_template(
    template_id: int,
    body: EmailTemplateUpdate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Edit an email template.
    - If user edits their OWN template (personal or forked): update in-place.
    - If user edits an ADMIN template and user is admin: update the original in-place.
    - If user edits an ADMIN template and user is NOT admin: create a fork (copy).
    """
    if not body.name or not body.html_body:
        raise HTTPException(status_code=400, detail="Name and HTML body are required")

    cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (template_id,))
    template = await cursor.fetchone()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    template = dict(template)

    now = datetime.utcnow().isoformat()

    # Case 1: User's own template (personal or fork) — edit in-place
    if template["owner_id"] is not None and template["owner_id"] == current_user["id"]:
        await db.execute(
            """UPDATE email_templates
               SET name = ?, subject = ?, html_body = ?, updated_at = ?
               WHERE id = ?""",
            (body.name, body.subject, body.html_body, now, template_id),
        )
        await db.commit()
        cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (template_id,))
        result = dict(await cursor.fetchone())
        result["is_default"] = False
        return result

    # Case 2: Admin template, edited by admin — update original in-place
    if template["owner_id"] is None and current_user["role"] == "admin":
        await db.execute(
            """UPDATE email_templates
               SET name = ?, subject = ?, html_body = ?, updated_at = ?
               WHERE id = ?""",
            (body.name, body.subject, body.html_body, now, template_id),
        )
        await db.commit()
        cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (template_id,))
        result = dict(await cursor.fetchone())
        result["is_default"] = True
        return result

    # Case 3: Admin template, edited by non-admin — create a fork
    if template["owner_id"] is None:
        # Check if user already has a fork of this template
        cursor = await db.execute(
            "SELECT * FROM email_templates WHERE owner_id = ? AND forked_from = ?",
            (current_user["id"], template_id),
        )
        existing_fork = await cursor.fetchone()

        if existing_fork:
            # Update the existing fork
            await db.execute(
                """UPDATE email_templates
                   SET name = ?, subject = ?, html_body = ?, updated_at = ?
                   WHERE id = ?""",
                (body.name, body.subject, body.html_body, now, existing_fork["id"]),
            )
            await db.commit()
            cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (existing_fork["id"],))
            result = dict(await cursor.fetchone())
            result["is_default"] = False
            return result
        else:
            # Create new fork
            cursor = await db.execute(
                """INSERT INTO email_templates (name, subject, html_body, owner_id, forked_from, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (body.name, body.subject, body.html_body, current_user["id"], template_id, now, now),
            )
            await db.commit()
            new_id = cursor.lastrowid
            cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (new_id,))
            result = dict(await cursor.fetchone())
            result["is_default"] = False
            return result

    # Fallback: user trying to edit someone else's template
    raise HTTPException(status_code=403, detail="Access denied")


@router.delete("/settings/email-templates/{template_id}")
async def delete_email_template(
    template_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Delete an email template.
    - User can delete only their own templates (owner_id = current_user.id).
    - User CANNOT delete admin/system templates (owner_id IS NULL) unless they are admin.
    - If user deletes a fork, the original admin template becomes visible again automatically.
    """
    cursor = await db.execute("SELECT * FROM email_templates WHERE id = ?", (template_id,))
    template = await cursor.fetchone()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    template = dict(template)

    # Admin template — only admin can delete
    if template["owner_id"] is None:
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Cannot delete default templates")
        await db.execute("DELETE FROM email_templates WHERE id = ?", (template_id,))
        await db.commit()
        return {"message": "Email Template deleted successfully"}

    # User's own template — allow delete
    if template["owner_id"] == current_user["id"]:
        await db.execute("DELETE FROM email_templates WHERE id = ?", (template_id,))
        await db.commit()
        return {"message": "Email Template deleted successfully"}

    # Someone else's template
    raise HTTPException(status_code=403, detail="Access denied")




# ──────────────────────────────────────────────
#  Settings: Google Calendar Config (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/google-calendar")
async def get_gcal_settings(admin: dict = Depends(require_admin)):
    """Get current Google Calendar configuration (admin only)."""
    return {
        "client_id": GCAL_CONFIG["client_id"],
        "client_secret": "***" if GCAL_CONFIG["client_secret"] else "",
        "refresh_token": "***" if GCAL_CONFIG["refresh_token"] else "",
        "calendar_email": GCAL_CONFIG["calendar_email"],
        "configured": is_gcal_configured(),
    }


@router.post("/settings/google-calendar")
async def save_gcal_settings(
    config: dict,
    admin: dict = Depends(require_admin),
):
    """Update Google Calendar configuration (admin only)."""
    # Don't overwrite masked values
    for key in ("client_secret", "refresh_token"):
        if config.get(key) == "***":
            del config[key]
    update_gcal_config(config)
    return {"message": "Google Calendar settings updated", "configured": is_gcal_configured()}


# ──────────────────────────────────────────────
#  Settings: Incoming Emails Gmail API Config (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/imap")
async def get_imap_settings(admin: dict = Depends(require_admin)):
    """Get current Incoming Email configuration (admin only)."""
    return {
        "client_id": GMAIL_FETCH_CONFIG["client_id"],
        "client_secret": "***" if GMAIL_FETCH_CONFIG["client_secret"] else "",
        "refresh_token": "***" if GMAIL_FETCH_CONFIG["refresh_token"] else "",
        "email": GMAIL_FETCH_CONFIG["email"],
        "configured": is_gmail_fetch_configured(),
    }


@router.post("/settings/imap")
async def save_imap_settings(
    config: dict,
    admin: dict = Depends(require_admin),
):
    """Update Incoming Email configuration (admin only)."""
    # Don't overwrite masked values
    for key in ("client_secret", "refresh_token"):
        if config.get(key) == "***":
            del config[key]
    update_gmail_fetch_config(config)
    return {"message": "Incoming email settings updated", "configured": is_gmail_fetch_configured()}


# ──────────────────────────────────────────────
#  Settings: Thank You Email Templates (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/thank-you-templates")
async def get_thank_you_templates(
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin)
):
    """Get current Thank You email templates (admin only)."""
    cursor = await db.execute("SELECT key, value FROM system_settings WHERE key IN ('posh_thank_you_template', 'contact_us_thank_you_template')")
    rows = await cursor.fetchall()
    settings = {row["key"]: row["value"] for row in rows}
    
    return {
        "posh": settings.get("posh_thank_you_template", POSH_DEFAULT_TEMPLATE),
        "contact_us": settings.get("contact_us_thank_you_template", CONTACT_US_DEFAULT_TEMPLATE)
    }


class ThankYouTemplatesUpdate(BaseModel):
    posh: str
    contact_us: str


@router.post("/settings/thank-you-templates")
async def save_thank_you_templates(
    body: ThankYouTemplatesUpdate,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin)
):
    """Update Thank You email templates (admin only)."""
    await db.execute(
        "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('posh_thank_you_template', ?)",
        (body.posh,)
    )
    await db.execute(
        "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('contact_us_thank_you_template', ?)",
        (body.contact_us,)
    )
    await db.commit()
    return {"message": "Thank You templates updated"}


@router.delete("/settings/thank-you-templates/{template_type}")
async def delete_thank_you_template(
    template_type: str,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin)
):
    """Delete a specific Thank You email template."""
    key = f"{template_type}_thank_you_template"
    if key not in ('posh_thank_you_template', 'contact_us_thank_you_template'):
        raise HTTPException(status_code=400, detail="Invalid template type")
        
    await db.execute("DELETE FROM system_settings WHERE key = ?", (key,))
    await db.commit()
    return {"message": f"{template_type} template deleted"}


# ──────────────────────────────────────────────
#  Settings: Admin CC Emails (Admin only)
# ──────────────────────────────────────────────
@router.get("/settings/cc-emails")
async def get_cc_emails(
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """Get all admin-managed default CC emails (admin only)."""
    cursor = await db.execute("SELECT * FROM cc_emails ORDER BY created_at DESC")
    rows = await cursor.fetchall()
    return [dict(row) for row in rows]


class CCEmailCreate(BaseModel):
    email: str
    scope_outgoing: Optional[bool] = True
    scope_meetings: Optional[bool] = True


@router.post("/settings/cc-emails")
async def add_cc_email(
    body: CCEmailCreate,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """Add a new admin-managed default CC email (admin only)."""
    from backend.validators import validate_email
    try:
        email = validate_email(body.email.strip().lower())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid email address")
    if not email:
        raise HTTPException(status_code=400, detail="Email cannot be empty")

    # Check for duplicates
    cursor = await db.execute("SELECT id FROM cc_emails WHERE email = ?", (email,))
    if await cursor.fetchone():
        raise HTTPException(status_code=400, detail="This CC email already exists")

    await db.execute(
        "INSERT INTO cc_emails (email, scope_outgoing, scope_meetings) VALUES (?, ?, ?)",
        (email, 1 if body.scope_outgoing else 0, 1 if body.scope_meetings else 0),
    )
    await db.commit()
    return {"message": f"CC email '{email}' added successfully"}


@router.delete("/settings/cc-emails/{cc_id}")
async def delete_cc_email(
    cc_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """Delete an admin-managed CC email (admin only)."""
    cursor = await db.execute("DELETE FROM cc_emails WHERE id = ?", (cc_id,))
    await db.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="CC email not found")
    return {"message": "CC email deleted successfully"}


# ──────────────────────────────────────────────
#  User CC Emails (per-user custom CCs)
# ──────────────────────────────────────────────
@router.get("/user-cc-emails")
async def get_user_cc_emails(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get the current user's custom CC emails."""
    cursor = await db.execute(
        "SELECT * FROM user_cc_emails WHERE user_id = ? ORDER BY created_at DESC",
        (current_user["id"],),
    )
    rows = await cursor.fetchall()
    return [dict(row) for row in rows]


class UserCCEmailCreate(BaseModel):
    email: str


@router.post("/user-cc-emails")
async def add_user_cc_email(
    body: UserCCEmailCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Add a new custom CC email for the current user."""
    from backend.validators import validate_email
    try:
        email = validate_email(body.email.strip().lower())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid email address")
    if not email:
        raise HTTPException(status_code=400, detail="Email cannot be empty")

    # Check for duplicates within user's own CCs
    cursor = await db.execute(
        "SELECT id FROM user_cc_emails WHERE user_id = ? AND email = ?",
        (current_user["id"], email),
    )
    if await cursor.fetchone():
        raise HTTPException(status_code=400, detail="This CC email already exists in your list")

    await db.execute(
        "INSERT INTO user_cc_emails (user_id, email) VALUES (?, ?)",
        (current_user["id"], email),
    )
    await db.commit()
    return {"message": f"CC email '{email}' added to your list"}


@router.delete("/user-cc-emails/{cc_id}")
async def delete_user_cc_email(
    cc_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Delete a user's own custom CC email."""
    cursor = await db.execute(
        "DELETE FROM user_cc_emails WHERE id = ? AND user_id = ?",
        (cc_id, current_user["id"]),
    )
    await db.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="CC email not found or access denied")
    return {"message": "CC email deleted successfully"}


# ──────────────────────────────────────────────
#  Public CC Emails Fetch (for campaign forms)
# ──────────────────────────────────────────────
@router.get("/cc-emails-for-send")
async def get_cc_emails_for_send(
    scope: str = "outgoing",
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Get CC emails for use in campaign/meeting forms.
    Returns admin defaults (filtered by scope) + current user's custom CCs.
    """
    # Admin defaults filtered by scope
    if scope == "meetings":
        cursor = await db.execute("SELECT id, email, scope_outgoing, scope_meetings FROM cc_emails WHERE scope_meetings = 1 ORDER BY created_at DESC")
    else:
        cursor = await db.execute("SELECT id, email, scope_outgoing, scope_meetings FROM cc_emails WHERE scope_outgoing = 1 ORDER BY created_at DESC")
    admin_rows = [dict(row) for row in await cursor.fetchall()]

    # User custom CCs
    cursor = await db.execute(
        "SELECT id, email FROM user_cc_emails WHERE user_id = ? ORDER BY created_at DESC",
        (current_user["id"],),
    )
    user_rows = [dict(row) for row in await cursor.fetchall()]

    return {"admin_cc": admin_rows, "user_cc": user_rows}


# ──────────────────────────────────────────────
#  Office Hours & Meeting Settings (Admin)
# ──────────────────────────────────────────────
@router.get("/settings/office-hours")
async def get_office_hours_settings(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get current office hours and meeting buffer settings."""
    return await get_office_hours(db)


@router.post("/settings/office-hours")
async def save_office_hours_settings(
    request: Request,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Update office hours and meeting buffer (admin only)."""
    body = await request.json()
    start_time = body.get("start_time", "09:00").strip()
    end_time = body.get("end_time", "18:00").strip()
    buffer_minutes = int(body.get("buffer_minutes", 30))

    # Normalize HH:MM:SS → HH:MM (browsers may include seconds)
    if len(start_time) > 5 and start_time[5:6] == ':':
        start_time = start_time[:5]
    if len(end_time) > 5 and end_time[5:6] == ':':
        end_time = end_time[:5]

    # Basic validation
    import re as _re
    if not _re.match(r'^\d{2}:\d{2}$', start_time):
        raise HTTPException(status_code=400, detail="Invalid start time format. Use HH:MM.")
    if not _re.match(r'^\d{2}:\d{2}$', end_time):
        raise HTTPException(status_code=400, detail="Invalid end time format. Use HH:MM.")
    if buffer_minutes < 0 or buffer_minutes > 120:
        raise HTTPException(status_code=400, detail="Buffer must be between 0 and 120 minutes.")

    # Ensure start < end
    start_h, start_m = map(int, start_time.split(":"))
    end_h, end_m = map(int, end_time.split(":"))
    if (start_h * 60 + start_m) >= (end_h * 60 + end_m):
        raise HTTPException(status_code=400, detail="Start time must be before end time.")

    await update_office_hours(db, start_time, end_time, buffer_minutes)
    return {"status": "saved", "start_time": start_time, "end_time": end_time, "buffer_minutes": buffer_minutes}


@router.get("/meeting-config")
async def get_meeting_config(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get meeting configuration for the scheduler UI (any authenticated user)."""
    office = await get_office_hours(db)
    return {
        "start_time": office["start_time"],
        "end_time": office["end_time"],
        "buffer_minutes": office["buffer_minutes"],
    }
