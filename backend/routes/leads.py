"""
Lead Routes — Full CRUD for leads, Excel upload, and OCR image scanning.
All endpoints are protected and scoped to the current user.
"""

import os
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
import aiosqlite

from backend.validators import validate_email, validate_phone

from backend.database import get_db
from backend.models.lead import LeadCreate, LeadResponse, LeadUpdate
from backend.middleware.auth_middleware import get_current_user
from backend.services.lead_service import (
    create_lead,
    get_leads_for_user,
    get_lead_by_id,
    update_lead,
    delete_lead,
    delete_all_leads,
    bulk_delete_leads,
    get_lead_stats,
    ensure_user_upload_dir,
    generate_booking_token,
    get_booking_token,
)

router = APIRouter(prefix="/api/leads", tags=["Leads"])


# ──────────────────────────────────────────────
#  Create a Lead Manually
# ──────────────────────────────────────────────
@router.post("", response_model=LeadResponse)
async def create_new_lead(
    lead_data: LeadCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Create a new lead manually."""
    try:
        lead = await create_lead(
            db=db,
            owner_id=current_user["id"],
            company_name=lead_data.company_name,
            contact_name=lead_data.contact_name,
            email=lead_data.email,
            phone=lead_data.phone,
            source=lead_data.source,
            notes=lead_data.notes,
        )
        return lead
    except ValueError as e:
        if "Duplicate lead" in str(e):
            raise HTTPException(status_code=400, detail=str(e))
        raise e


# ──────────────────────────────────────────────
#  List All Leads (with optional filters)
# ──────────────────────────────────────────────
@router.get("")
async def list_leads(
    status_filter: str = Query(None, alias="status"),
    source: str = Query(None),
    search: str = Query(None),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    List leads. Salesperson sees their own; Admin sees all.
    Optional filters: ?status=new&search=acme&source=Website
    """
    leads = await get_leads_for_user(
        db=db,
        user_id=current_user["id"],
        role=current_user["role"],
        status=status_filter,
        search=search,
        source=source,
    )
    return leads


# ──────────────────────────────────────────────
#  Get Lead Statistics (for dashboard)
# ──────────────────────────────────────────────
@router.get("/stats")
async def lead_stats(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get lead statistics for the dashboard."""
    stats = await get_lead_stats(db, current_user["id"], current_user["role"])
    return stats


# ──────────────────────────────────────────────
#  Delete All Leads
# ──────────────────────────────────────────────
@router.delete("/all")
async def delete_all_existing_leads(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Delete all leads and their associated campaigns for the user."""
    count = await delete_all_leads(db, current_user["id"], current_user["role"])
    return {"message": f"Deleted {count} leads successfully"}


# ──────────────────────────────────────────────
#  Bulk Delete Selected Leads (fast)
# ──────────────────────────────────────────────
@router.post("/bulk-delete")
async def bulk_delete_leads_endpoint(
    payload: dict,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Delete multiple leads and their associated data in a single fast transaction.
    Expects JSON body: { "lead_ids": [1, 2, 3, ...] }
    """
    lead_ids = payload.get("lead_ids", [])
    if not lead_ids or not isinstance(lead_ids, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="lead_ids must be a non-empty list",
        )
    if len(lead_ids) > 10000:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete more than 10,000 leads at once",
        )

    count = await bulk_delete_leads(
        db=db,
        lead_ids=lead_ids,
        user_id=current_user["id"],
        role=current_user["role"],
    )
    return {"message": f"Deleted {count} leads successfully", "deleted": count}


# ──────────────────────────────────────────────
#  Download Sample Excel Format
# ──────────────────────────────────────────────
@router.get("/sample-format")
async def download_sample_format():
    """
    Download a sample Excel file for bulk uploading leads.
    """
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Leads"
    
    # Headers exactly matching expected mappings
    headers = ["Company Name", "Contact Person Name", "Email", "Phone"]
    ws.append(headers)
    
    # Add a sample row to guide the user
    sample_row = ["Acme Corp", "John Doe", "john.doe@example.com", "1234567890"]
    ws.append(sample_row)

    # Auto-adjust column widths for better UX
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="sample_leads_format.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# ──────────────────────────────────────────────
#  Download Leads as Excel
# ──────────────────────────────────────────────
@router.get("/download")
async def download_leads_excel(
    status_filter: str = Query(None, alias="status"),
    source: str = Query(None),
    search: str = Query(None),
    date: str = Query(None),
    limit: int = Query(0),
    offset: int = Query(0),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Download leads as a structured Excel file.
    Columns: Company Name, Contact Person Name, Email, Phone, Sent By, Date.
    Supports all existing filters plus date, limit, and offset for pagination.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from datetime import datetime as dt

    # Build query with JOIN to get owner's full_name as "Sent By"
    query = """
        SELECT l.company_name, l.contact_name, l.email, l.phone,
               u.full_name as sent_by, l.created_at
        FROM leads l
        LEFT JOIN users u ON l.owner_id = u.id
    """
    params = []
    conditions = []

    # Ownership filter
    if current_user["role"] != "admin":
        conditions.append("l.owner_id = ?")
        params.append(current_user["id"])

    if status_filter:
        conditions.append("l.status = ?")
        params.append(status_filter)

    if source:
        conditions.append("l.source = ?")
        params.append(source)

    if search:
        conditions.append(
            "(l.company_name LIKE ? OR l.contact_name LIKE ? OR l.email LIKE ? OR l.phone LIKE ?)"
        )
        search_term = f"%{search}%"
        params.extend([search_term, search_term, search_term, search_term])

    if date:
        conditions.append("DATE(l.created_at) = ?")
        params.append(date)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY l.created_at DESC"

    if limit and limit > 0:
        query += " LIMIT ? OFFSET ?"
        params.extend([limit, offset])

    cursor = await db.execute(query, params)
    rows = await cursor.fetchall()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers_list = ["Company Name", "Contact Person Name", "Email", "Phone", "Sent By", "Date"]
    ws.append(headers_list)

    # Apply header styling
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data styling
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center")
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for row_idx, row in enumerate(rows, start=2):
        row_dict = dict(row)
        # Format date
        date_val = ""
        if row_dict.get("created_at"):
            try:
                parsed_date = dt.fromisoformat(row_dict["created_at"].replace("Z", "+00:00"))
                date_val = parsed_date.strftime("%d-%b-%Y")
            except (ValueError, AttributeError):
                date_val = str(row_dict["created_at"])[:10]

        data_row = [
            row_dict.get("company_name") or "",
            row_dict.get("contact_name") or "",
            row_dict.get("email") or "",
            row_dict.get("phone") or "",
            row_dict.get("sent_by") or "",
            date_val,
        ]
        ws.append(data_row)

        # Apply data styling with alternating row colors
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    today = dt.now().strftime("%Y-%m-%d")
    resp_headers = {
        "Content-Disposition": f'attachment; filename="leads_export_{today}.xlsx"'
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


# ──────────────────────────────────────────────
#  Get a Single Lead
# ──────────────────────────────────────────────
@router.get("/{lead_id}", response_model=LeadResponse)
async def get_single_lead(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get a single lead by ID."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # Ensure user can only access their own leads (unless admin)
    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    return lead


# ──────────────────────────────────────────────
#  Update a Lead
# ──────────────────────────────────────────────
@router.put("/{lead_id}", response_model=LeadResponse)
async def update_existing_lead(
    lead_id: int,
    lead_data: LeadUpdate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Update an existing lead."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        updated = await update_lead(
            db=db,
            lead_id=lead_id,
            company_name=lead_data.company_name,
            contact_name=lead_data.contact_name,
            email=lead_data.email,
            phone=lead_data.phone,
            notes=lead_data.notes,
            status=lead_data.status,
        )
    except ValueError as e:
        if "Duplicate lead" in str(e):
            raise HTTPException(status_code=400, detail=str(e))
        raise e
    return updated


# ──────────────────────────────────────────────
#  Delete a Lead
# ──────────────────────────────────────────────
@router.delete("/{lead_id}")
async def delete_existing_lead(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Delete a lead and its associated campaigns."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    await delete_lead(db, lead_id)
    return {"message": "Lead deleted successfully"}


# ──────────────────────────────────────────────
#  Upload Excel File (Bulk Import)
# ──────────────────────────────────────────────
@router.post("/upload/excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Upload an Excel file to bulk-import leads.
    
    Expected columns (case-insensitive):
      company_name | contact_name | email | phone | notes
    
    At least one of company_name/contact_name/email/phone must be present.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an .xlsx, .xls, or .csv file.",
        )

    try:
        import openpyxl
        import csv
        import xlrd

        # Save file temporarily to user's upload directory
        user_dir = ensure_user_upload_dir(current_user["id"])
        file_path = os.path.join(user_dir, file.filename)

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        try:
            headers = []
            rows_data = []

            if file.filename.endswith('.csv'):
                # Parse CSV file directly from memory
                text_content = content.decode('utf-8-sig', errors='replace')
                
                # Robust Delimiter Sniffing
                try:
                    dialect = csv.Sniffer().sniff(text_content[:2048])
                    reader = csv.reader(io.StringIO(text_content), dialect)
                except csv.Error:
                    # Fallback if sniffing fails
                    if ";" in text_content[:1024] and "," not in text_content[:1024]:
                        reader = csv.reader(io.StringIO(text_content), delimiter=";")
                    elif "\t" in text_content[:1024]:
                        reader = csv.reader(io.StringIO(text_content), delimiter="\t")
                    else:
                        reader = csv.reader(io.StringIO(text_content))

                try:
                    raw_headers = next(reader)
                    
                    # If standard comma reader failed and gave us 1 column containing semicolons, fix it
                    if len(raw_headers) == 1 and ";" in raw_headers[0]:
                        reader = csv.reader(io.StringIO(text_content), delimiter=";")
                        raw_headers = next(reader)

                    for h in raw_headers:
                        val = str(h).strip().lower().replace(" ", "_") if h else ""
                        headers.append(val)
                    for row in reader:
                        rows_data.append(row)
                except StopIteration:
                    pass
            elif file.filename.endswith('.xls'):
                # Parse older Excel file using xlrd
                book = xlrd.open_workbook(file_contents=content)
                if book.nsheets > 0:
                    sheet = book.sheet_by_index(0)
                    if sheet.nrows > 0:
                        raw_headers = sheet.row_values(0)
                        for h in raw_headers:
                            val = str(h).strip().lower().replace(" ", "_") if h else ""
                            headers.append(val)
                        for row_idx in range(1, sheet.nrows):
                            rows_data.append(sheet.row_values(row_idx))
            else:
                # Parse newer Excel file (.xlsx) directly from memory
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active

                # Get headers from first row
                for cell in ws[1]:
                    val = str(cell.value).strip().lower().replace(" ", "_") if cell.value else ""
                    headers.append(val)

                # Get rows
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows_data.append(row)

            # Map columns
            col_map = {}
            for i, h in enumerate(headers):
                if "company" in h:
                    col_map["company_name"] = i
                elif "contact" in h or "name" in h:
                    col_map["contact_name"] = i
                elif "email" in h or "mail" in h:
                    col_map["email"] = i
                elif "phone" in h or "mobile" in h or "tel" in h:
                    col_map["phone"] = i
                elif "note" in h or "remark" in h:
                    col_map["notes"] = i

            # Validate that we didn't just merge everything into a single column
            if len(headers) <= 1 and (";" in "".join(headers) or "\t" in "".join(headers) or len("".join(headers)) > 30):
                raise HTTPException(
                    status_code=400,
                    detail="Invalid file structure detected. All columns were merged into one. Please ensure your CSV uses commas or download and use the Sample Format."
                )

            # Check if at least one recognizable column was found
            if not col_map:
                raise HTTPException(
                    status_code=400,
                    detail="Could not map any columns. Please ensure your headers include Company, Contact, Email, or Phone, or download the Sample Format."
                )

            # Import rows
            imported = 0
            skipped = 0
            errors = []

            for row_idx, row in enumerate(rows_data, start=2):
                try:
                    def get_val(col_name):
                        if col_name in col_map and col_map[col_name] < len(row):
                            val = row[col_map[col_name]]
                            if val is not None and str(val).strip() != "":
                                return str(val).strip()
                        return None

                    company = get_val("company_name")
                    contact = get_val("contact_name")
                    email = get_val("email")
                    phone = get_val("phone")
                    notes = get_val("notes")

                    # Skip completely empty rows
                    if not any([company, contact, email, phone]):
                        skipped += 1
                        continue

                    # Clean up "None" strings
                    if company == "None": company = None
                    if contact == "None": contact = None
                    if email == "None": email = None
                    if phone == "None": phone = None
                    if notes == "None": notes = None

                    # Require email
                    if not email:
                        errors.append(f"Row {row_idx}: Email is required")
                        skipped += 1
                        continue

                    # Validate email and phone formats
                    try:
                        email = validate_email(email)
                        if phone:
                            phone = validate_phone(phone)
                    except ValueError as ve:
                        errors.append(f"Row {row_idx}: {str(ve)}")
                        skipped += 1
                        continue

                    await create_lead(
                        db=db,
                        owner_id=current_user["id"],
                        company_name=company,
                        contact_name=contact,
                        email=email,
                        phone=phone,
                        source="excel",
                        notes=notes,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    skipped += 1

            return {
                "message": f"Import complete: {imported} leads imported, {skipped} skipped",
                "imported": imported,
                "skipped": skipped,
                "errors": errors[:10],  # Return max 10 errors
                "file": file.filename,
            }

        finally:
            # Always delete the uploaded file after processing to prevent disk buildup
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass  # Non-critical cleanup

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel processing failed: {str(e)}")



# ──────────────────────────────────────────────
#  Upload Visiting Card Image (OCR)
# ──────────────────────────────────────────────
@router.post("/upload/ocr")
async def upload_ocr(
    file: UploadFile = File(...),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Upload a visiting card image and extract contact data using OCR.
    Supported formats: jpg, jpeg, png, bmp, tiff
    
    Returns the extracted data for review before saving.
    """
    # Validate file type
    allowed_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif')
    if not file.filename.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Supported: {', '.join(allowed_extensions)}",
        )

    try:
        from backend.services.ocr_service import extract_text_from_image, parse_business_card

        # Save file temporarily to user's upload directory
        user_dir = ensure_user_upload_dir(current_user["id"])
        file_path = os.path.join(user_dir, f"card_{file.filename}")

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        try:
            # Run OCR
            raw_text = extract_text_from_image(file_path)

            # Parse structured data from OCR text
            parsed = parse_business_card(raw_text)

            return {
                "message": "Card scanned successfully",
                "extracted": parsed,
                "file": file.filename,
            }

        finally:
            # Always delete the uploaded image after processing to prevent disk buildup
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass  # Non-critical cleanup

    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")



# ──────────────────────────────────────────────
#  Save OCR Result as Lead
# ──────────────────────────────────────────────
@router.post("/ocr/save", response_model=LeadResponse)
async def save_ocr_lead(
    lead_data: LeadCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Save OCR-extracted data as a new lead.
    Called after the user reviews and edits the extracted data.
    """
    try:
        lead = await create_lead(
            db=db,
            owner_id=current_user["id"],
            company_name=lead_data.company_name,
            contact_name=lead_data.contact_name,
            email=lead_data.email,
            phone=lead_data.phone,
            source="ocr",
            notes=lead_data.notes,
        )
        return lead
    except ValueError as e:
        if "Duplicate lead" in str(e):
            raise HTTPException(status_code=400, detail=str(e))
        raise e


# ──────────────────────────────────────────────
#  Generate / Get Booking Token for a Lead
# ──────────────────────────────────────────────
@router.post("/{lead_id}/booking-token")
async def create_booking_token(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Generate a booking token for a lead. Returns the booking URL."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        token = await generate_booking_token(db, lead_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    return {
        "token": token,
        "lead_id": lead_id,
    }


@router.get("/{lead_id}/booking-token")
async def get_lead_booking_token(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get the existing booking token for a lead (generates one if missing)."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        result = await get_booking_token(db, lead_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    return {
        "token": result["token"],
        "used": result["used"],
        "lead_id": lead_id,
    }

